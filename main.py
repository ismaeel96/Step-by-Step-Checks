import tkinter
import tkinter.messagebox
import datetime

from tkinter import ttk
from tkinter import *
from openpyxl import Workbook
from openpyxl import load_workbook
from os.path import exists
from pathlib import Path

Excel_File_Path = 'MetalChecks.xlsx'
path = Path(Excel_File_Path)
print("Created by: Ismaeel Jamil \nPhone: 586-258-6567")
print("")

if path.is_file():
    print(f'The file {Excel_File_Path} exists')
    workbook = load_workbook(filename="MetalChecks.xlsx")
else:
    print(f'The file {Excel_File_Path} does not exist')
    workbook = Workbook()
    print(f'Created file:{Excel_File_Path}')
sheet = workbook.active

Start_Column = 1
Check_Start_Row = 4
Num_Of_Checks = 0
Row_Start_Index = Check_Start_Row

while sheet.cell(row = 2, column = Start_Column).value != None:
  #print(sheet.cell(row = 2, column = Start_Column).value)
  #print("column: ", Start_Column)
  Start_Column += 1

while sheet.cell(row = Check_Start_Row+Num_Of_Checks, column = 1).value != None:
  #print(sheet.cell(row = Check_Start_Row+Num_Of_Checks, column = 1).value)
  #print("Row: ",Check_Start_Row+Num_Of_Checks)
  Num_Of_Checks += 1

#print("starting column: ",Start_Column)
#print("Number of Checks: ",Num_Of_Checks)

def OK_Command():
    global Row_Start_Index, Check_Start_Row, Start_Column,Num_Of_Checks,Check_Label,OK_Button,NOK_Button
    #print("OK")
    cell = sheet.cell(row=Row_Start_Index, column=Start_Column-1)
    cell.value = "OK"
    workbook.save(filename="MetalChecks.xlsx")
    Row_Start_Index += 1

    if Row_Start_Index >= Num_Of_Checks+Check_Start_Row:
        Row_Start_Index = Check_Start_Row
        SN_Input.delete(0, END)
        SN_Input.focus_set()
        OK_Button.grid_remove()
        NOK_Button.grid_remove()

    Check_Label.configure(text = sheet.cell(Row_Start_Index, 1).value)
def NOK_Command():
    global Row_Start_Index, Check_Start_Row, Start_Column,Num_Of_Checks,Check_Label,OK_Button,NOK_Button
    #print("NOK")
    cell = sheet.cell(row=Row_Start_Index, column=Start_Column-1)
    cell.value = "NOK"
    workbook.save(filename="MetalChecks.xlsx")
    Row_Start_Index += 1

    if Row_Start_Index >= Num_Of_Checks+Check_Start_Row:
        Row_Start_Index = Check_Start_Row
        SN_Input.delete(0, END)
        SN_Input.focus_set()
        OK_Button.grid_remove()
        NOK_Button.grid_remove()

    Check_Label.configure(text=sheet.cell(Row_Start_Index, 1).value)
def SN_Command(SN, Input_Col, Input_Row):
    global Start_Column
    #print(SN, "(",Input_Col," , ",Input_Row,")")
    cell = sheet.cell(row=Input_Row, column=Input_Col)
    cell.value = SN
    cell_date = sheet.cell(row=Input_Row+1, column=Input_Col)
    cell_date.value = str(datetime.datetime.now())
    workbook.save(filename="MetalChecks.xlsx")
    Start_Column += 1



def make_invisible(widget):
    widget.pack_forget()


window = Tk()
window.title("Frame Quality Check - Ismaeel Jamil")
frm = ttk.Frame(window, padding=10)
#window.geometry("600x400")
frm.grid()
Barcode_label = ttk.Label(frm, text="Scan Barcode: ")
Barcode_label.grid(column=0, row=0)


SN_Input=ttk.Entry(frm, width = 35)
SN_Input.grid(column=1, row=0)
SN_Input.focus_set()

OK_Button = ttk.Button(frm, text="OK", command=OK_Command, width=25)
NOK_Button = ttk.Button(frm, text="NOK", command=NOK_Command, width=25)
OK_Button.grid_remove()
NOK_Button.grid_remove()

Check_Label = ttk.Label(frm, text=sheet.cell(Row_Start_Index, 1).value, wraplength=250, justify="center")
Check_Label.grid(column=0, row=2)

def return_event(event):
    #print("You hit return.")
    condition = str(window.focus_get())
    #print(condition)
    if condition == ".!frame.!entry":
        #print("if passed")
        SN_Command(str(SN_Input.get()), Start_Column, 2)

        OK_Button.grid(column=1, row=4)
        NOK_Button.grid(column=2, row=4)


        Barcode_label.focus_set()
        #print(window.focus_get())

SN_Input.bind('<Return>', return_event)


window.mainloop()

#top = tkinter.Tk()
#def helloCallBack():
#   tkinter.messagebox.showinfo( "Hello Python", "Hello World")

#OK = tkinter.Button(top, text ="OK", command = helloCallBack)
#NOK = tkinter.Button(top, text ="NOK", command = helloCallBack)

#OK.pack()
#NOK.pack()

#top.mainloop()
# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print('End program')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
